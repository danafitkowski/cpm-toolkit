"""Builds '3-Week Lookahead Starter Template.xlsx': a blank, generic field lookahead."""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
NAVY = "14213D"
TEAL = "0F766E"
LIGHT = "F5F7FA"
BLUE_INPUT = "0000FF"

wb = Workbook()

# ---- Instructions sheet ----
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 90
ins["A1"] = "3-Week Lookahead Starter Template"
ins["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
lines = [
    "",
    "How to use this file:",
    "1. Go to the '3-Week Lookahead' tab.",
    "2. Fill in the 'Week of' date at the top.",
    "3. Blue text marks cells meant to be edited. Replace the example row (row 6) with your own activities, then delete it or keep it as a reference and add rows below.",
    "4. The daily columns under WEEK 1 are for a quick status mark per day (e.g. an 'x', a percentage, or a short note). Use whatever convention your team already uses.",
    "5. WEEK 2 and WEEK 3 columns are for the planned activity for that week at a glance, not a daily breakdown.",
    "6. 'Days Remaining' calculates automatically from the Planned Finish date, so leave that column's formula alone.",
    "",
    "This is a manual, P6-independent template. Nothing here reads or writes P6 files.",
]
for i, line in enumerate(lines, start=2):
    c = ins.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, size=11, bold=line.startswith("How to use"))
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ---- Main lookahead sheet ----
ws = wb.create_sheet("3-Week Lookahead")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F1")
ws["A1"] = "3-WEEK LOOKAHEAD"
ws["A1"].font = Font(name=FONT, size=18, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws["G1"] = "Week of:"
ws["G1"].font = Font(name=FONT, size=11, bold=True)
ws["G1"].alignment = Alignment(horizontal="right", vertical="center")
ws["G1"].fill = PatternFill("solid", fgColor=NAVY)
ws["G1"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
ws.merge_cells("H1:I1")
ws["H1"] = "FILL IN"
ws["H1"].font = Font(name=FONT, size=11, bold=True, color=BLUE_INPUT)
ws["H1"].fill = PatternFill("solid", fgColor=NAVY)
ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
ws["H1"].number_format = "mmm d, yyyy"

headers_row = 3
group_row = 3
sub_row = 4

cols = [
    ("A", "Act. ID", 10),
    ("B", "Activity Description", 32),
    ("C", "Area / Location", 16),
    ("D", "Responsible", 16),
    ("E", "Planned Start", 13),
    ("F", "Planned Finish", 13),
    ("G", "Mon", 6), ("H", "Tue", 6), ("I", "Wed", 6), ("J", "Thu", 6), ("K", "Fri", 6), ("L", "Sat", 6),
    ("M", "Week 2 Plan", 20),
    ("N", "Week 3 Plan", 20),
    ("O", "Predecessor / Constraint", 24),
    ("P", "Days Remaining", 14),
    ("Q", "Notes", 28),
]
for letter, _, width in cols:
    ws.column_dimensions[letter].width = width

thin = Side(style="thin", color="DDE3EA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Group header row 3
ws.merge_cells("A3:A4"); ws["A3"] = "Act. ID"
ws.merge_cells("B3:B4"); ws["B3"] = "Activity Description"
ws.merge_cells("C3:C4"); ws["C3"] = "Area / Location"
ws.merge_cells("D3:D4"); ws["D3"] = "Responsible"
ws.merge_cells("E3:E4"); ws["E3"] = "Planned Start"
ws.merge_cells("F3:F4"); ws["F3"] = "Planned Finish"
ws.merge_cells("G3:L3"); ws["G3"] = "WEEK 1 (THIS WEEK)"
ws["G4"] = "Mon"; ws["H4"] = "Tue"; ws["I4"] = "Wed"; ws["J4"] = "Thu"; ws["K4"] = "Fri"; ws["L4"] = "Sat"
ws.merge_cells("M3:M4"); ws["M3"] = "Week 2 Plan"
ws.merge_cells("N3:N4"); ws["N3"] = "Week 3 Plan"
ws.merge_cells("O3:O4"); ws["O3"] = "Predecessor / Constraint"
ws.merge_cells("P3:P4"); ws["P3"] = "Days Remaining"
ws.merge_cells("Q3:Q4"); ws["Q3"] = "Notes"

for row in (3, 4):
    for col in range(1, 18):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
ws.row_dimensions[3].height = 20
ws.row_dimensions[4].height = 20

# Example row (row 6), clearly marked
example_note_row = 5
ws.cell(row=5, column=1, value="EXAMPLE ROW: replace with your own data, or delete")
ws.cell(row=5, column=1).font = Font(name=FONT, size=9, italic=True, color="B45309")
ws.merge_cells("A5:Q5")

ex_row = 6
example = ["A1010", "Form and pour footings - Grid A-C", "Area 2", "J. Smith",
           date(2026, 8, 3), date(2026, 8, 7), "x", "x", "x", "", "", "",
           "Strip forms, begin backfill", "Set anchor bolts", "Concrete cure 3 days (FS+3)",
           None, "Inspector booked for Fri AM"]
for i, val in enumerate(example, start=1):
    cell = ws.cell(row=ex_row, column=i, value=val)
    cell.font = Font(name=FONT, size=10, color=BLUE_INPUT)
    cell.border = border
    cell.alignment = Alignment(vertical="center", wrap_text=(i in (2, 13, 14, 17)))
# Days Remaining formula (column P = 16)
ws.cell(row=ex_row, column=16, value="=F6-TODAY()")
ws.cell(row=ex_row, column=16).font = Font(name=FONT, size=10, color="000000")
ws.cell(row=ex_row, column=16).number_format = "0"
ws.cell(row=ex_row, column=16).border = border
ws.cell(row=ex_row, column=5).number_format = "mmm d, yyyy"
ws.cell(row=ex_row, column=6).number_format = "mmm d, yyyy"

# 24 blank rows ready to fill in, each with the Days Remaining formula pre-loaded
for r in range(7, 31):
    for c in range(1, 18):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.font = Font(name=FONT, size=10)
        if (r % 2) == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.cell(row=r, column=5).number_format = "mmm d, yyyy"
    ws.cell(row=r, column=6).number_format = "mmm d, yyyy"
    p = ws.cell(row=r, column=16, value=f"=F{r}-TODAY()")
    p.number_format = "0"

ws.freeze_panes = "A5"

# ---- Professional polish ----
wb.properties.creator = "CPM Toolkit"
wb.properties.title = "3-Week Lookahead Starter Template"
wb.properties.subject = "Blank, editable 3-week look-ahead schedule template"
wb.properties.company = "CPM Toolkit"

foot_row = 32
ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=17)
foot = ws.cell(row=foot_row, column=1, value="CPM Toolkit  |  3-Week Lookahead Starter Template  |  v1.0")
foot.font = Font(name=FONT, size=8, italic=True, color="33415C")
for c in range(1, 18):
    ws.cell(row=foot_row, column=c).border = Border(top=Side(style="thin", color="DDE3EA"))

ws.sheet_properties.tabColor = TEAL
ins.sheet_properties.tabColor = NAVY

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "3:4"
ws.page_margins.left = 0.35
ws.page_margins.right = 0.35
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.print_options.horizontalCentered = True

wb.save("3-Week Lookahead Starter Template.xlsx")
print("saved")
