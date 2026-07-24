"""Builds 'Schedule Risk & Contingency Starter Kit.xlsx': PERT 3-point risk register.

Honestly labeled as 3-point (PERT) estimating, not a full Monte Carlo simulation.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
NAVY = "14213D"
TEAL = "0F766E"
LIGHT = "F5F7FA"
BLUE_INPUT = "0000FF"
thin = Side(style="thin", color="DDE3EA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 92
ins["A1"] = "Schedule Risk & Contingency Starter Kit"
ins["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
lines = [
    "",
    "What this is:",
    "A 3-point (PERT) estimating workbook. For each activity or risk item, enter an Optimistic, Most Likely, "
    "and Pessimistic duration. The sheet calculates a PERT-weighted expected duration and a standard deviation "
    "per item, then combines them into a project-level expected duration and a P50 / P80 / P90 range.",
    "",
    "What this is not:",
    "This is not a full Monte Carlo simulation. It doesn't model correlations between activities or run thousands "
    "of iterations. It's the standard lightweight alternative: fast to build, transparent, and good enough to put "
    "a number on contingency for most schedules. For a full simulation on a complex or high-value schedule, that's "
    "still worth doing separately.",
    "",
    "How to use it:",
    "1. Go to the 'Risk Register' tab.",
    "2. Blue text marks cells meant to be edited. Row 5 is an example, replace it with your own items.",
    "3. List either individual activities or risk-bearing groups of activities, whichever gives you a more "
    "manageable list. 15-30 items is a normal range.",
    "4. The bottom summary combines every row into a project-level Expected, P80, and P90 duration.",
    "5. '% of Total Variance' ranks which items are driving the uncertainty; the top few are worth managing "
    "first.",
]
for i, line in enumerate(lines, start=2):
    c = ins.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, size=11, bold=line.endswith(":"))
    c.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("Risk Register")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:H1")
ws["A1"] = "SCHEDULE RISK & CONTINGENCY (3-POINT / PERT)"
ws["A1"].font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

cols = [
    ("A", "Item / Activity", 30),
    ("B", "Optimistic (O)", 12),
    ("C", "Most Likely (M)", 12),
    ("D", "Pessimistic (P)", 12),
    ("E", "PERT Expected (Te)", 14),
    ("F", "Std Dev (SD)", 12),
    ("G", "Variance", 10),
    ("H", "% of Total Variance", 14),
]
for letter, _, width in cols:
    ws.column_dimensions[letter].width = width

header_row = 3
for i, (letter, label, _) in enumerate(cols, start=1):
    cell = ws.cell(row=header_row, column=i, value=label)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[header_row].height = 28

ws.cell(row=4, column=1, value="EXAMPLE ROW: replace with your own items. Durations in working days.")
ws.cell(row=4, column=1).font = Font(name=FONT, size=9, italic=True, color="B45309")
ws.merge_cells("A4:H4")

ex_row = 5
n_blank = 19
data_last_row = ex_row + n_blank  # 24

def add_row(r, name, o, m, p, editable):
    color = BLUE_INPUT if editable else None
    font_name = Font(name=FONT, size=10, color=color) if editable else Font(name=FONT, size=10)
    ws.cell(row=r, column=1, value=name).font = font_name
    for col, val in ((2, o), (3, m), (4, p)):
        c = ws.cell(row=r, column=col, value=val)
        c.font = font_name
        c.alignment = Alignment(horizontal="center")
    te = ws.cell(row=r, column=5, value=f"=IF(AND(B{r}<>\"\",C{r}<>\"\",D{r}<>\"\"),(B{r}+4*C{r}+D{r})/6,\"\")")
    te.font = Font(name=FONT, size=10)
    te.number_format = "0.0"
    te.alignment = Alignment(horizontal="center")
    sd = ws.cell(row=r, column=6, value=f"=IF(AND(B{r}<>\"\",D{r}<>\"\"),(D{r}-B{r})/6,\"\")")
    sd.font = Font(name=FONT, size=10)
    sd.number_format = "0.0"
    sd.alignment = Alignment(horizontal="center")
    var = ws.cell(row=r, column=7, value=f"=IF(F{r}<>\"\",F{r}^2,\"\")")
    var.font = Font(name=FONT, size=10)
    var.number_format = "0.00"
    var.alignment = Alignment(horizontal="center")
    pct = ws.cell(row=r, column=8,
                   value=f"=IF(AND(G{r}<>\"\",$G${data_last_row + 2}>0),G{r}/$G${data_last_row + 2},\"\")")
    pct.font = Font(name=FONT, size=10)
    pct.number_format = "0.0%"
    pct.alignment = Alignment(horizontal="center")
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = border

add_row(ex_row, "Excavation - unexpected groundwater", 5, 8, 20, True)
for r in range(ex_row + 1, data_last_row + 1):
    add_row(r, None, None, None, None, False)
    if (r % 2) == 0:
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT)

sum_row = data_last_row + 2
ws.cell(row=sum_row, column=1, value="Sum of variance (used above):").font = Font(name=FONT, size=9, italic=True, color="33415C")
ws.cell(row=sum_row, column=7, value=f"=SUM(G{ex_row}:G{data_last_row})").font = Font(name=FONT, size=9, italic=True)
ws.cell(row=sum_row, column=7).number_format = "0.00"

label_row = sum_row + 2
ws.cell(row=label_row, column=1, value="PROJECT SUMMARY").font = Font(name=FONT, size=12, bold=True, color=NAVY)

r1 = label_row + 1
ws.cell(row=r1, column=1, value="Expected duration (P50), sum of Te:").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r1, column=3, value=f"=SUM(E{ex_row}:E{data_last_row})").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r1, column=3).number_format = "0.0"

r2 = r1 + 1
ws.cell(row=r2, column=1, value="Combined std dev (independent items):").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r2, column=3, value=f"=SQRT(SUM(G{ex_row}:G{data_last_row}))").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r2, column=3).number_format = "0.0"

r3 = r2 + 1
ws.cell(row=r3, column=1, value="P80 duration (P50 + 0.84 SD):").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r3, column=3, value=f"=C{r1}+0.84*C{r2}").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r3, column=3).number_format = "0.0"

r4 = r3 + 1
ws.cell(row=r4, column=1, value="P90 duration (P50 + 1.28 SD):").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r4, column=3, value=f"=C{r1}+1.28*C{r2}").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r4, column=3).number_format = "0.0"

r5 = r4 + 1
ws.cell(row=r5, column=1, value="Suggested contingency (P80 minus P50):").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r5, column=3, value=f"=C{r3}-C{r1}").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=r5, column=3).number_format = "0.0"

ws.freeze_panes = "A5"

# ---- Professional polish ----
wb.properties.creator = "CPM Toolkit"
wb.properties.title = "Schedule Risk & Contingency Starter Kit"
wb.properties.subject = "3-point (PERT) risk register with P50/P80/P90 contingency"
wb.properties.company = "CPM Toolkit"

foot_row = r5 + 2
ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=8)
foot = ws.cell(row=foot_row, column=1, value="CPM Toolkit  |  Schedule Risk & Contingency Starter Kit  |  v1.0")
foot.font = Font(name=FONT, size=8, italic=True, color="33415C")

ws.sheet_properties.tabColor = TEAL
ins.sheet_properties.tabColor = NAVY

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = f"{header_row}:{header_row}"
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.print_options.horizontalCentered = True

wb.save("Schedule Risk & Contingency Starter Kit.xlsx")
print("saved, data rows", ex_row, "..", data_last_row, "summary at", label_row)
