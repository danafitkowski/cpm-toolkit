"""Builds 'Resource Histogram Starter.xlsx': weekly trade loading grid + stacked chart."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

FONT = "Arial"
NAVY = "14213D"
TEAL = "0F766E"
LIGHT = "F5F7FA"
BLUE_INPUT = "0000FF"
thin = Side(style="thin", color="DDE3EA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 90
ins["A1"] = "Resource Histogram Starter"
ins["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
lines = [
    "",
    "How to use this file:",
    "1. Go to the 'Histogram' tab.",
    "2. Blue text marks cells meant to be edited. Rows 6 to 7 are examples, replace them with your own trades.",
    "3. Enter a headcount or hours figure per week for each trade in the grid.",
    "4. Peak and Average calculate automatically per row. The Total row sums every trade per week.",
    "5. The chart below the grid updates automatically as you edit the numbers.",
    "6. Need more than 12 weeks? Select column M, copy it, and insert copies before the Peak column, then extend the Total row and chart range.",
    "7. The chart only charts the 2 example rows to start, so it isn't cluttered with blank series. Right-click the chart, choose 'Select Data', and extend the series range to cover your own trade rows once you've filled them in.",
]
for i, line in enumerate(lines, start=2):
    c = ins.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, size=11, bold=line.startswith("How to use"))
    c.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("Histogram")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F1")
ws["A1"] = "RESOURCE HISTOGRAM"
ws["A1"].font = Font(name=FONT, size=18, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws["G1"] = "Week 1 starting:"
ws.merge_cells("G1:H1")
ws["G1"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
ws["G1"].fill = PatternFill("solid", fgColor=NAVY)
ws["G1"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
ws.merge_cells("I1:J1")
ws["I1"] = "FILL IN"
ws["I1"].font = Font(name=FONT, size=11, bold=True, color=BLUE_INPUT)
ws["I1"].fill = PatternFill("solid", fgColor=NAVY)
ws["I1"].alignment = Alignment(horizontal="center", vertical="center")
ws["I1"].number_format = "mmm d, yyyy"

N_WEEKS = 12
header_row = 3
ws.column_dimensions["A"].width = 24
week_cols = []
for w in range(N_WEEKS):
    col = 2 + w  # B..M
    week_cols.append(col)
    ws.column_dimensions[ws.cell(row=header_row, column=col).column_letter].width = 8.5
peak_col = 2 + N_WEEKS       # N
avg_col = peak_col + 1       # O
ws.column_dimensions[ws.cell(row=header_row, column=peak_col).column_letter].width = 9
ws.column_dimensions[ws.cell(row=header_row, column=avg_col).column_letter].width = 9

ws.cell(row=header_row, column=1, value="Trade / Resource")
for w, col in enumerate(week_cols, start=1):
    ws.cell(row=header_row, column=col, value=f"Wk {w}")
ws.cell(row=header_row, column=peak_col, value="Peak")
ws.cell(row=header_row, column=avg_col, value="Average")
for col in range(1, avg_col + 1):
    cell = ws.cell(row=header_row, column=col)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
ws.row_dimensions[header_row].height = 18

note_row = 4
ws.cell(row=note_row, column=1, value="EXAMPLE ROWS 5-6: replace with your own trades")
ws.cell(row=note_row, column=1).font = Font(name=FONT, size=9, italic=True, color="B45309")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=avg_col)

examples = [
    ("Carpenters", [4, 6, 8, 8, 8, 8, 6, 6, 4, 4, 2, 0]),
    ("Electricians", [0, 0, 2, 4, 4, 6, 6, 6, 6, 4, 2, 2]),
]
r = 5
for name, vals in examples:
    ws.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10, color=BLUE_INPUT)
    ws.cell(row=r, column=1).border = border
    for col, v in zip(week_cols, vals):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name=FONT, size=10, color=BLUE_INPUT)
        c.alignment = Alignment(horizontal="center")
        c.border = border
    first_l = ws.cell(row=r, column=week_cols[0]).column_letter
    last_l = ws.cell(row=r, column=week_cols[-1]).column_letter
    pk = ws.cell(row=r, column=peak_col, value=f"=MAX({first_l}{r}:{last_l}{r})")
    pk.border = border; pk.alignment = Alignment(horizontal="center")
    av = ws.cell(row=r, column=avg_col, value=f"=AVERAGE({first_l}{r}:{last_l}{r})")
    av.border = border; av.alignment = Alignment(horizontal="center"); av.number_format = "0.0"
    r += 1

blank_start = r
for r in range(blank_start, blank_start + 10):
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=1).font = Font(name=FONT, size=10)
    for col in week_cols:
        c = ws.cell(row=r, column=col)
        c.border = border
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(horizontal="center")
    first_l = ws.cell(row=r, column=week_cols[0]).column_letter
    last_l = ws.cell(row=r, column=week_cols[-1]).column_letter
    pk = ws.cell(row=r, column=peak_col, value=f"=MAX({first_l}{r}:{last_l}{r})")
    pk.border = border; pk.alignment = Alignment(horizontal="center")
    av = ws.cell(row=r, column=avg_col, value=f"=AVERAGE({first_l}{r}:{last_l}{r})")
    av.border = border; av.alignment = Alignment(horizontal="center"); av.number_format = "0.0"
    if (r % 2) == 0:
        for col in range(1, avg_col + 1):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)

data_last_row = blank_start + 9
total_row = data_last_row + 1
ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=total_row, column=1).border = border
ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="E5EEF0")
for col in week_cols:
    letter = ws.cell(row=total_row, column=col).column_letter
    c = ws.cell(row=total_row, column=col, value=f"=SUM({letter}5:{letter}{data_last_row})")
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.border = border
    c.fill = PatternFill("solid", fgColor="E5EEF0")
peak_l = ws.cell(row=total_row, column=peak_col).column_letter
ws.cell(row=total_row, column=peak_col,
        value=f"=MAX({ws.cell(row=total_row, column=week_cols[0]).column_letter}{total_row}:"
              f"{ws.cell(row=total_row, column=week_cols[-1]).column_letter}{total_row})")
for c in (peak_col, avg_col):
    cell = ws.cell(row=total_row, column=c)
    cell.font = Font(name=FONT, size=10, bold=True)
    cell.border = border
    cell.fill = PatternFill("solid", fgColor="E5EEF0")
    cell.alignment = Alignment(horizontal="center")

ws.freeze_panes = "B5"

# Stacked column chart of the trade rows (not the Total row) across the 12 weeks
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
chart.overlap = 100
chart.title = "Weekly headcount by trade"
chart.y_axis.title = "Headcount"
chart.x_axis.title = "Week"
chart.height = 9
chart.width = 24

cats = Reference(ws, min_col=week_cols[0], max_col=week_cols[-1], min_row=header_row, max_row=header_row)
# Charted against the 2 example rows only, so the chart isn't cluttered with
# empty-named series before the buyer fills anything in. Instructions tab
# explains how to extend the series range once real trades are added.
data = Reference(ws, min_col=1, max_col=week_cols[-1], min_row=5, max_row=6)
chart.add_data(data, titles_from_data=True, from_rows=True)
chart.set_categories(cats)
ws.add_chart(chart, f"A{total_row + 3}")

# ---- Professional polish ----
wb.properties.creator = "CPM Toolkit"
wb.properties.title = "Resource Histogram Starter"
wb.properties.subject = "Weekly trade-loading grid with built-in stacked chart"
wb.properties.company = "CPM Toolkit"

foot_row = total_row + 1
ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=avg_col)
foot = ws.cell(row=foot_row, column=1, value="CPM Toolkit  |  Resource Histogram Starter  |  v1.0")
foot.font = Font(name=FONT, size=8, italic=True, color="33415C")

ws.sheet_properties.tabColor = TEAL
ins.sheet_properties.tabColor = NAVY

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = f"{header_row}:{header_row}"
ws.page_margins.left = 0.35
ws.page_margins.right = 0.35
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.print_options.horizontalCentered = True

wb.save("Resource Histogram Starter.xlsx")
print("saved, data rows 5..", data_last_row, "total row", total_row)
