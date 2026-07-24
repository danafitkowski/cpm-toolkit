"""Builds 'Schedule Health Checklist.xlsx': a manual, DCMA-14-style walkthrough checklist."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
NAVY = "14213D"
TEAL = "0F766E"
LIGHT = "F5F7FA"

wb = Workbook()
ws = wb.active
ws.title = "Checklist"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:E1")
ws["A1"] = "SCHEDULE HEALTH CHECKLIST"
ws["A1"].font = Font(name=FONT, size=18, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:E2")
ws["A2"] = "Mark each item Pass / Flag / N-A as you review. Status column has a dropdown."
ws["A2"].font = Font(name=FONT, size=10, italic=True, color="33415C")
ws.row_dimensions[2].height = 18

headers = ["#", "Check", "What to look for", "Status", "Notes"]
widths = [5, 30, 52, 12, 30]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    col = ws.cell(row=4, column=i, value=h)
    col.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    col.fill = PatternFill("solid", fgColor=TEAL)
    col.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[chr(64 + i)].width = w
ws.row_dimensions[4].height = 20

items = [
    ("Logic completeness", "Every activity has at least one predecessor and one successor, except the true start and finish milestones."),
    ("Leads (negative lag)", "No relationship uses a negative lag to fake an overlap that isn't real logic."),
    ("Excessive lag", "Long positive lags are actual waiting time (cure, cool-down, approval), not a placeholder for missing logic."),
    ("Relationship types", "The large majority of relationships are Finish-to-Start. Start-to-Start and Finish-to-Finish pairs are used deliberately, not as a shortcut."),
    ("Hard constraints", "Mandatory or 'must finish on' constraints are rare and each one is justified (contract date, regulatory date, etc.), not used to force a date."),
    ("High float", "Activities carrying unusually high total float are checked: is that float real, or is a missing successor hiding it?"),
    ("Negative float", "Nothing on the schedule is showing negative float without an explanation already in hand."),
    ("Long-duration activities", "No single activity runs longer than about 44 working days without being broken into trackable pieces."),
    ("Invalid or out-of-window dates", "Actual dates make sense against status (no actual start on a not-started activity; no missing actual finish on a completed one)."),
    ("Data date currency", "The schedule's data date is current for this reporting period, not stale from a prior update."),
    ("Out-of-sequence progress", "Activities aren't shown progressing before their predecessors finish without a documented reason."),
    ("Critical path sanity", "The critical path tells a story that makes sense to someone who knows the project, not a path that jumps around illogically."),
    ("Calendar assignment", "Activities are on the calendar that matches how the work actually happens (5-day, 6-day, weather, cure, etc.)."),
    ("Milestone naming", "Contract and key milestones are named and coded consistently so they're easy to find and report on."),
]

r = 5
for i, (check, note) in enumerate(items, start=1):
    ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value=check).font = Font(name=FONT, size=10, bold=True)
    c3 = ws.cell(row=r, column=3, value=note)
    c3.font = Font(name=FONT, size=10)
    c3.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=4, value="")
    ws.cell(row=r, column=5, value="").alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = Border(*(Side(style="thin", color="DDE3EA"),) * 4)
        if r % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT)
        if c in (1, 4):
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 34
    r += 1

last_row = r - 1
dv = DataValidation(type="list", formula1='"Pass,Flag,N/A"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"D5:D{last_row}")

# Summary
sum_row = last_row + 2
ws.cell(row=sum_row, column=2, value="Passed:").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=sum_row, column=3, value=f'=COUNTIF(D5:D{last_row},"Pass")&" of "&COUNTA(D5:D{last_row})&" checked"').font = Font(name=FONT, size=10)
ws.cell(row=sum_row + 1, column=2, value="Flagged:").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=sum_row + 1, column=3, value=f'=COUNTIF(D5:D{last_row},"Flag")').font = Font(name=FONT, size=10)

wb.save("Schedule Health Checklist.xlsx")
print("saved, last_row", last_row)
