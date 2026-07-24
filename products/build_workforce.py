"""Builds 'Workforce Plan Template.xlsx': mobilize/peak/demobilize planning by trade."""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

FONT = "Arial"
NAVY = "14213D"
TEAL = "0F766E"
LIGHT = "F5F7FA"
BLUE_INPUT = "0000FF"
thin = Side(style="thin", color="DDE3EA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 90
ins["A1"] = "Workforce Plan Template"
ins["A1"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
lines = [
    "",
    "How to use this file:",
    "1. Go to the 'Workforce Plan' tab.",
    "2. Blue text marks cells meant to be edited. Row 5 is an example, replace it with your own crews.",
    "3. 'Weeks on Site' and the summary totals at the bottom calculate automatically from your dates.",
    "4. This is a planning snapshot (peak headcount and mobilization dates per trade), not a weekly loading grid.",
    "   For week-by-week headcount over time, use the companion Resource Histogram Starter instead.",
    "5. The chart only plots the example row to start. Right-click it, choose 'Select Data', and extend the range to your own rows once they're filled in.",
]
for i, line in enumerate(lines, start=2):
    c = ins.cell(row=i, column=1, value=line)
    c.font = Font(name=FONT, size=11, bold=line.startswith("How to use"))
    c.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("Workforce Plan")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:G1")
ws["A1"] = "WORKFORCE PLAN"
ws["A1"].font = Font(name=FONT, size=18, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

cols = [
    ("A", "Trade / Crew", 20),
    ("B", "Peak Headcount", 14),
    ("C", "Mobilize", 13),
    ("D", "Peak By", 13),
    ("E", "Demobilize", 13),
    ("F", "Weeks on Site", 13),
    ("G", "Lead Time to Hire (wks)", 16),
    ("H", "Notes", 30),
]
for letter, _, width in cols:
    ws.column_dimensions[letter].width = width

header_row = 3
for i, (letter, label, _) in enumerate(cols, start=1):
    cell = ws.cell(row=header_row, column=i, value=label)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[header_row].height = 28

ws.cell(row=4, column=1, value="EXAMPLE ROW: replace with your own crews")
ws.cell(row=4, column=1).font = Font(name=FONT, size=9, italic=True, color="B45309")
ws.merge_cells("A4:H4")

ex_row = 5
example = ["Electricians", 6, date(2026, 9, 1), date(2026, 10, 15), date(2026, 12, 1), None, 3,
           "Local union hall, 3-week call lead time"]
for i, val in enumerate(example, start=1):
    cell = ws.cell(row=ex_row, column=i, value=val)
    cell.font = Font(name=FONT, size=10, color=BLUE_INPUT)
    cell.border = border
    cell.alignment = Alignment(vertical="center", wrap_text=(i == 8))
ws.cell(row=ex_row, column=3).number_format = "mmm d, yyyy"
ws.cell(row=ex_row, column=4).number_format = "mmm d, yyyy"
ws.cell(row=ex_row, column=5).number_format = "mmm d, yyyy"
wk = ws.cell(row=ex_row, column=6, value=f"=(E{ex_row}-C{ex_row})/7")
wk.number_format = "0.0"
wk.font = Font(name=FONT, size=10)
wk.border = border

blank_start = 6
n_blank = 12
for r in range(blank_start, blank_start + n_blank):
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.font = Font(name=FONT, size=10)
        if c in (3, 4, 5):
            cell.number_format = "mmm d, yyyy"
        if (r % 2) == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT)
    wkf = ws.cell(row=r, column=6, value=f"=IF(AND(E{r}<>\"\",C{r}<>\"\"),(E{r}-C{r})/7,\"\")")
    wkf.number_format = "0.0"

data_last_row = blank_start + n_blank - 1  # 17
sum_row = data_last_row + 2
ws.cell(row=sum_row, column=1, value="Total peak headcount:").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=sum_row, column=2, value=f"=SUM(B{ex_row}:B{data_last_row})").font = Font(name=FONT, size=10, bold=True)
ws.cell(row=sum_row + 1, column=1, value="Earliest mobilization:").font = Font(name=FONT, size=10, bold=True)
mob = ws.cell(row=sum_row + 1, column=2, value=f"=MIN(C{ex_row}:C{data_last_row})")
mob.font = Font(name=FONT, size=10, bold=True); mob.number_format = "mmm d, yyyy"
ws.cell(row=sum_row + 2, column=1, value="Latest demobilization:").font = Font(name=FONT, size=10, bold=True)
demob = ws.cell(row=sum_row + 2, column=2, value=f"=MAX(E{ex_row}:E{data_last_row})")
demob.font = Font(name=FONT, size=10, bold=True); demob.number_format = "mmm d, yyyy"

ws.freeze_panes = "A5"

chart = BarChart()
chart.type = "col"
chart.title = "Peak headcount by trade"
chart.y_axis.title = "Peak headcount"
chart.height = 8
chart.width = 18
cats = Reference(ws, min_col=1, max_col=1, min_row=ex_row, max_row=ex_row)
data = Reference(ws, min_col=2, max_col=2, min_row=header_row, max_row=ex_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, f"A{sum_row + 4}")

wb.save("Workforce Plan Template.xlsx")
print("saved, data rows", ex_row, "..", data_last_row, "sum row", sum_row)
